# !/usr/bin/env python
# -*- coding:utf-8
# author:mlh
# datetime:2019/6/5 下午2:02

import time
import unittest
from init import InitTest
from common import login
from selenium.webdriver.common.action_chains import ActionChains
from log import logTest
import random
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import openpyxl


def code():
    Acode = openpyxl.load_workbook("..\\data\\Astocks.xlsx")
    sheets = Acode.sheetnames
    sheet = Acode[sheets[0]]
    # print(sheet)
    num = random.randint(2, sheet.max_row+1)
    # print(num)
    code = "".join([str(sheet.cell(row=num, column=1).value)])
    # print(code)
    return code


def result(self, t):
    time.sleep(3)
    r_div = self.driver.find_elements(By.CSS_SELECTOR, ".s-relational-pos")[t]
    r_ul = r_div.find_element(By.XPATH, "ul")
    r_li = r_ul.find_elements(By.XPATH, "li")
    for i in range(-len(r_li)+1, 1):
        r_a = r_li[i].find_element(By.XPATH, "a")
        time.sleep(3)
        # print(i)
        r_a.click()
        time.sleep(5)
        r_div = self.driver.find_elements(By.CSS_SELECTOR, ".s-relational-pos")[t]
        r_ul = r_div.find_element(By.XPATH, "ul")


class Chart(InitTest):
    '''测试图谱'''
    def test_001_Chart(self):
        '''打开图谱页面'''
        # self.driver.implicitly_wait(60)
        self.driver.get("http://www.ibdata.cn/#!/chart")
        login.login(self)
        # self.driver.find_element_by_link_text("图谱").click()
        # time.sleep(15)
        self.assertEqual(self.driver.current_url, "http://www.ibdata.cn/#!/chart")
        logTest.log("当前是图谱页面")

    def test_002_Chart(self):
        '''输入要查询的公司'''
        so = self.driver.find_element_by_class_name("ep-input")
        self.assertTrue(so.is_enabled())  # 验证输入框可编辑
        # suv_list = ["00", "600", "300", "002", "200", "900", "87", "43", "83"]
        # num = random.randint(0, len(suv_list) - 1)
        # so.send_keys(suv_list[num])
        # ul = self.driver.find_element_by_class_name("company-menu")
        # lis = ul.find_elements_by_xpath("li")
        # num1 = random.randint(0, len(lis) - 1)
        # time.sleep(3)
        # print(num1)
        # lis[num1].click()
        inp = code()
        so.send_keys(inp)
        time.sleep(3)
        logTest.log("已经输入要查询的公司")
        ActionChains(self.driver).move_by_offset(0, 0).click().perform()
        time.sleep(3)
        btn = self.driver.find_element_by_class_name("btn-search")
        btn.click()
        time.sleep(60)
        logTest.log("点击查询")
        ActionChains(self.driver).move_by_offset(0, 0).click().perform()

    def test_003_Chart(self):
        '''对结果图的操作'''
        result(self, 0)
        result(self, 1)
        btn_del = self.driver.find_element(By.CSS_SELECTOR, ".glyphicon.glyphicon-remove")
        btn_del.click()

    def test_004_Chart(self):
        '''同名相连和上下游的操作'''
        # 同名相连
        time.sleep(5)
        checkbox = self.driver.find_element_by_css_selector(".el-checkbox-inner")
        checkbox.click()
        time.sleep(5)
        logTest.log("同名相连")
        checkbox.click()
        time.sleep(5)
        logTest.log("取消同名相连")
        # 上下游

        top_icon = self.driver.find_element_by_css_selector(".Supply-top-icon")
        top_icon.click()
        time.sleep(30)
        logTest.log("上游")

        try:
            self.driver.switch_to.default_content()
            bt_del = self.driver.find_element(By.CSS_SELECTOR, ".btn.btn-del.right30")
            bt_del.click()
            logTest.log("没有上游公司")
        except:
            logTest.log("有上游")

        bot_icon = self.driver.find_element_by_class_name("Supply-bot-icon")
        bot_icon.click()
        time.sleep(30)
        logTest.log("下游")
        try:
            self.driver.switch_to.default_content()
            bn = self.driver.find_element(By.CSS_SELECTOR, ".btn.btn-del.right30")
            bn.click()
            logTest.log("没有下游公司")
        except:
            logTest.log("有下游")

    # def test_005_Chart(self):
    #     '''自定义操作'''
    #     # 自定义信息
    #
    #     btn_rule = self.driver.find_element_by_class_name("btn-rule")
    #     btn_rule.click()
    #     self.driver.switch_to.default_content()
    #     time.sleep(3)
    #
    #     label = self.driver.find_element(By.XPATH, "//*[@id='dialog']/div/div[2]/div/ul/li[5]/div[2]/div/input")
    #     ActionChains(self.driver).click(label).perform()
    #     # label.click()
    #     time.sleep(15)
    #     new_built = self.driver.find_element(By.CSS_SELECTOR, ".new-built.right")
    #     new_built.click()
    #     time.sleep(10)
    #     logTest.log("自定义操作")

    def test_006_Chart(self):
        '''添加可比公司'''
        btn_search = self.driver.find_element(By.CLASS_NAME, "btn-search")
        btn_search.send_keys(Keys.ENTER)
        time.sleep(60)
        im = self.driver.find_elements(By.CLASS_NAME, "node")[0]
        child = im.find_elements(By.XPATH, "./*")[2]
        child.click()
        self.driver.switch_to.default_content()
        btn_add = self.driver.find_element(By.CLASS_NAME, "btn-add")
        so = self.driver.find_element(By.CLASS_NAME, "other-comp-input")
        # suv_list = ["00", "600", "300", "002", "200", "900", "87", "43", "83"]
        # num = random.randint(0, len(suv_list) - 1)
        # so.send_keys(suv_list[num])
        # ul = self.driver.find_element_by_class_name("company-menu")
        # lis = ul.find_elements_by_xpath("li")
        # num1 = random.randint(0, len(lis) - 1)
        # time.sleep(3)
        # # print(num1)
        # lis[num1].click()
        # time.sleep(3)
        inp1 = code()
        so.send_keys(inp1)
        btn_add.click()
        logTest.log("添加可比公司一个")
        time.sleep(60)
        im = self.driver.find_elements(By.CLASS_NAME, "node")[0]
        child = im.find_elements(By.XPATH, "./*")[2]
        child.click()
        self.driver.switch_to.default_content()
        btn_add1 = self.driver.find_element(By.CLASS_NAME, "btn-add")
        btn_add1.click()
        logTest.log("不更改输入，直接点击添加")
        time.sleep(60)
        self.driver.switch_to.default_content()
        btn_btn = self.driver.find_element(By.CSS_SELECTOR, ".btn.btn-del.right30")
        btn_btn.click()
        logTest.log("重复添加")
        time.sleep(5)
        im = self.driver.find_elements(By.CLASS_NAME, "node")[0]
        child = im.find_elements(By.XPATH, "./*")[2]
        child.click()
        time.sleep(15)
        self.driver.switch_to.default_content()
        delete_icon = self.driver.find_element(By.CLASS_NAME, "delete-icon")
        delete_icon.click()
        time.sleep(5)
        logTest.log("取消添加可比公司")

    def test_007_Chart(self):
        '''切换为人名，并进行搜索'''
        names = self.driver.find_elements(By.CSS_SELECTOR, ".el-radio-inner")
        names[0].click()
        Name = openpyxl.load_workbook("..\\data\\name_map.xlsx")
        sheets = Name.sheetnames
        sheet = Name[sheets[0]]
        # print(sheet)
        num = random.randint(2, sheet.max_row + 1)
        # print(num)
        name = "".join([str(sheet.cell(row=num, column=1).value)])
        so = self.driver.find_element(By.CLASS_NAME, "ep-input")
        so.clear()
        time.sleep(3)
        so.send_keys(name)
        btn = self.driver.find_element(By.CLASS_NAME, "btn-search")
        btn.click()
        time.sleep(60)


if __name__ == "__main__":
    unittest.main(verbosity=2)


