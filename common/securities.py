# !/user/bin/env python
# -*- coding:utf-8 -*-
# author：mlh

import openpyxl
import random


def Asecurities():
    '''A股证券代码'''
    Asecurities = openpyxl.load_workbook("..\\data\\Astocks.xlsx")
    sheets = Asecurities.sheetnames
    sheet = Asecurities[sheets[0]]
    # print(sheet)
    num = random.randint(2, sheet.max_row+1)
    # print(num)
    securitiesA = "".join([str(sheet.cell(row=num, column=2).value)])
    return securitiesA


def N3Bsecurities():
    '''新三板证券简称'''
    N3Bsecurities = openpyxl.load_workbook("..\\data\\N3B.xlsx")
    sheets1 = N3Bsecurities.sheetnames
    sheet = N3Bsecurities[sheets1[0]]
    # print(sheet)
    num = random.randint(2, sheet.max_row + 1)
    # print(num)
    securitiesN3B = "".join([str(sheet.cell(row=num, column=2).value)])
    return securitiesN3B


def Bondsecurities():
    '''债券证券简称'''
    Bondsecurities = openpyxl.load_workbook("..\\data\\Bond.xlsx")
    sheets1 = Bondsecurities.sheetnames
    sheet = Bondsecurities[sheets1[0]]
    # print(sheet)
    num = random.randint(2, sheet.max_row + 1)
    # print(num)
    securitiesBond = "".join([str(sheet.cell(row=num, column=1).value)])
    return securitiesBond
