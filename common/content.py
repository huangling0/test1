# !/user/bin/env python
# -*- coding:utf-8 -*-
# author：mlh

import openpyxl
import random


def content():
	we = openpyxl.load_workbook("..\\data\\input.xlsx")
	sheets = we.sheetnames
	sheet = we[sheets[0]]
	num = random.randint(2, sheet.max_row+1)
	content = "".join([str(sheet.cell(row=num, column=1).value)])
	return content



