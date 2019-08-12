# !/user/bin/env python
# -*- coding:utf-8 -*-
# author：mlh

import openpyxl
import random


# def financial():
# 	we = openpyxl.load_workbook("..\\data\\financial.xlsx")
# 	sheets = we.sheetnames
# 	print(sheets)
# 	for i in range(len(sheets)):
# 		sheet = we[sheets[i]]
# 		print(sheet)
# 		for j in range(1, sheet.max_column+1):
# 			t = "".join([str(sheet.cell(row=1, column=j).value)])
#
# 			print(t)
# 			for n in range(1, sheet.max_column+1):
# 				data = "".join([str(sheet.cell(row=n, column=j).value)])
# 				print(data)
#
#
# financial()


def financial(i, j):
	we = openpyxl.load_workbook("..\\data\\financial.xlsx")
	sheets = we.sheetnames
	sheet = we[sheets[i]]
	# print(sheet)
	# t = "".join([str(sheet.cell(row=1, column=j).value)])
	# print(t)
	num = random.randint(2, sheet.max_row + 1)
	data = "".join([str(sheet.cell(row=num, column=j).value)])
	# print(data)

	return data


def hkfinancial(i, j):
	we = openpyxl.load_workbook("..\\data\\HKStocks.xlsx")
	sheets = we.sheetnames
	sheet = we[sheets[i]]
	num = random.randint(2, sheet.max_row + 1)
	data = "".join([str(sheet.cell(row=num, column=j).value)])
	return data

