# !/user/bin/env python
# -*- coding:utf-8 -*-
# author：mlh

import openpyxl
import random


def Acode():
    '''A股证券代码'''
    Acode = openpyxl.load_workbook("..\\data\\Astocks.xlsx")
    sheets = Acode.sheetnames
    sheet = Acode[sheets[0]]
    # print(sheet)
    num = random.randint(2, sheet.max_row+1)
    # print(num)
    codeA = "".join([str(sheet.cell(row=num, column=3).value)])
    return codeA


def N3Bcode():
    '''新三板证券代码'''
    N3Bcode = openpyxl.load_workbook("..\\data\\N3B.xlsx")
    sheets1 = N3Bcode.sheetnames
    sheet = N3Bcode[sheets1[0]]
    # print(sheet)
    num = random.randint(2, sheet.max_row + 1)
    # print(num)
    codeN3B = "".join([str(sheet.cell(row=num, column=1).value)])
    return codeN3B


def Hcode():
    '''港股证券代码'''
    Hcode = openpyxl.load_workbook("..\\data\\HKStocks.xlsx")
    sheets1 = Hcode.sheetnames
    sheet = Hcode[sheets1[0]]
    # print(sheet)
    num = random.randint(2, sheet.max_row + 1)
    # print(num)
    codeH = "".join([str(sheet.cell(row=num, column=2).value)])
    return codeH


def Bondcode():
    '''债券证券代码'''
    Bondcode = openpyxl.load_workbook("..\\data\\Bond.xlsx")
    sheets1 = Bondcode.sheetnames
    sheet = Bondcode[sheets1[0]]
    # print(sheet)
    num = random.randint(2, sheet.max_row + 1)
    # print(num)
    codeBond = "".join([str(sheet.cell(row=num, column=2).value)])
    return codeBond

