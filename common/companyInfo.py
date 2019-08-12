# !/user/bin/env python
# -*- coding:utf-8 -*-
# author：mlh

import openpyxl
import random


def N3BCompanyInfo(i):
	'''
	:param i:
	i == 1:公司全程
	i == 2：证券简称
	i == 3：证券代码
	i == 4：法人
	i == 5：注册地
	i == 6：所属地
	:return:
	'''
	Info = openpyxl.load_workbook("..\\data\\N3B_company.xlsx")
	sheets = Info.sheetnames
	sheet = Info[sheets[0]]
	num = random.randint(2, sheet.max_row)
	company_info = " ".join([str(sheet.cell(row=num, column=i).value)])

	return company_info


def AcompanyInfo(i):
	'''
	:param i:
	i == 1: 公司全程
	i == 2：证券简称
	i == 3：证券代码
	i == 4：注册地
	i == 5：行业
	i == 6：注册资本
	i == 7：市值
	i == 8：市盈率
	i == 9：全场质押比率
	:return:
	'''
	Info = openpyxl.load_workbook("..\\data\\Astocks.xlsx")
	sheets = Info.sheetnames
	sheet = Info[sheets[0]]
	num = random.randint(2, sheet.max_row)
	company_info = " ".join([str(sheet.cell(row=num, column=i).value)])
	# print(company_info)
	return company_info


def HKcompanyInfo(i):
	'''
	:param i:
	i == 1: 公司全程
	i == 2：证券简称
	i == 3：证券代码
	i == 4：注册地
	:return:
	'''
	Info = openpyxl.load_workbook("..\\data\\HKStocks.xlsx")
	sheets = Info.sheetnames
	sheet = Info[sheets[0]]
	num = random.randint(2, sheet.max_row)
	company_info = " ".join([str(sheet.cell(row=num, column=i).value)])
	# print(company_info)
	return company_info



